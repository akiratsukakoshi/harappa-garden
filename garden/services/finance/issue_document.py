#!/usr/bin/env python3
"""issue_document.py — finance 区画 Mode G(発行)の道具。

確定済みの構造正本(md frontmatter)を入力に、社印・書式・数式を保った雛形
.xlsx へ値だけ差し込んで、見積書(d)・請求書(f)の発行物を出力する。

設計の前提(ベンダー中立):
  - 知識・型・確定値はすべて md(soil の案件台帳 + finance/templates の様式)に残る
  - この道具は openpyxl だけで動く。LLM にも特定 SaaS にも依存しない
  - PDF 化は手作業(ガクチョが Excel で開いて捺印 → PDF)

入力 md の frontmatter(soil の finance/templates/*.md 形式に準拠):
  doc:        御見積書 | 御請求書        # 必須。どちらの雛形に差し込むか
  client:     宛先(社名)               # 必須
  subject:    件名
  # --- 見積書 ---
  estimate_date / valid_until
  # --- 請求書 ---
  invoice_date / due_date
  # --- 共通 ---
  line_items:
    - name: 品名
      qty: 数量
      unit_price: 単価(税抜)

  ※ 日付を省くと雛形の =TODAY() / 既定値をそのまま残す。
  ※ 小計・消費税・合計は雛形の数式が再計算する(ここでは触らない)。

使い方:
  python3 issue_document.py path/to/請求書.md [--out path.xlsx] [--dry-run]
"""

import sys
import argparse
import datetime
from pathlib import Path

import yaml
import openpyxl
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
TEMPLATES = HERE / "templates"

# doc 種別 → 雛形ファイル。
# 社印(丸印)は雛形 .xlsx に画像として埋め込まれており、Pillow が入っていれば
# openpyxl が load→save で保持する(requirements の Pillow は社印保持のため必須。
# 無いと openpyxl は画像を読めず save 時に落とす)。
TEMPLATE_BY_DOC = {
    "御見積書": TEMPLATES / "estimate-template.xlsx",
    "御請求書": TEMPLATES / "invoice-template.xlsx",
}

# 明細エリア(雛形の構造に対応)
INVOICE_ITEM_START, INVOICE_ITEM_END = 19, 40   # A19:F40(最大 22 行)
ESTIMATE_ITEM_START, ESTIMATE_ITEM_END = 16, 31  # B16:F31(最大 16 行)


def parse_frontmatter(md_path: Path) -> dict:
    """md の先頭 --- ... --- ブロックを YAML として読む。"""
    text = md_path.read_text(encoding="utf-8")
    if not text.startswith("---"):
        raise ValueError(f"{md_path} に frontmatter(--- ブロック)がありません")
    parts = text.split("---", 2)
    if len(parts) < 3:
        raise ValueError(f"{md_path} の frontmatter が閉じていません")
    fm = yaml.safe_load(parts[1]) or {}
    return fm


def to_date(v):
    """'YYYY-MM-DD' 文字列 / date / datetime を datetime.date に正規化。"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return datetime.date.fromisoformat(str(v).strip())


# 和暦の表示書式(値は日付のまま、表示だけ「令和8年7月25日」にする)
WAREKI_DATE_FMT = '[$-411]ggge"年"m"月"d"日"'


def set_date(ws, coord, value):
    """日付セルに date を入れ、和暦書式を適用する。値が空ならセルは触らない。"""
    d = to_date(value)
    if d is not None:
        ws[coord] = d
        ws[coord].number_format = WAREKI_DATE_FMT


def _check_items(items, limit, kind):
    if not items:
        raise ValueError(f"line_items が空です({kind})")
    if len(items) > limit:
        raise ValueError(
            f"明細が {len(items)} 行 = 雛形の上限 {limit} 行を超えます({kind})。"
            "明細をまとめるか、雛形を拡張してください。"
        )


def fill_invoice(ws, fm: dict):
    set_date(ws, "F3", fm.get("invoice_date"))
    ws["A4"] = f"{fm['client']} 御中"
    if fm.get("subject"):
        ws["A9"] = fm["subject"]
        base = ws["A9"].font
        ws["A9"].font = Font(name=base.name, size=base.size or 11, bold=True)
        ws.row_dimensions[9].height = 24   # 件名行を広げる

    # 明細エリアをクリア(数式ごと一旦消す)
    for r in range(INVOICE_ITEM_START, INVOICE_ITEM_END + 1):
        for col in "ABCDEF":
            ws[f"{col}{r}"] = None

    items = fm.get("line_items", [])
    _check_items(items, INVOICE_ITEM_END - INVOICE_ITEM_START + 1, "請求書")
    for i, it in enumerate(items):
        r = INVOICE_ITEM_START + i
        ws[f"A{r}"] = it["name"]
        ws[f"B{r}"] = it["unit_price"]
        ws[f"C{r}"] = it.get("qty", 1)
        ws[f"D{r}"] = f"=B{r}*C{r}"          # 小計
        ws[f"E{r}"] = f"=D{r}*0.1"           # 消費税額
        ws[f"F{r}"] = f"=D{r}+E{r}"          # 税込小計

    set_date(ws, "F44", fm.get("due_date"))


def fill_estimate(ws, fm: dict):
    set_date(ws, "F3", fm.get("estimate_date"))
    ws["B4"] = fm["client"]
    if fm.get("subject"):
        ws["B7"] = fm["subject"]
    set_date(ws, "C10", fm.get("valid_until"))

    # 明細エリアをクリア(品名 B は B:C 結合 → 左上 B に書く)
    for r in range(ESTIMATE_ITEM_START, ESTIMATE_ITEM_END + 1):
        for col in ("B", "D", "E", "F"):
            ws[f"{col}{r}"] = None

    items = fm.get("line_items", [])
    _check_items(items, ESTIMATE_ITEM_END - ESTIMATE_ITEM_START + 1, "見積書")
    for i, it in enumerate(items):
        r = ESTIMATE_ITEM_START + i
        ws[f"B{r}"] = it["name"]
        ws[f"D{r}"] = it.get("qty", 1)
        ws[f"E{r}"] = it["unit_price"]
        ws[f"F{r}"] = f"=D{r}*E{r}"           # 金額(税抜)


def issue(md_path: Path, out_path: Path = None, dry_run: bool = False) -> Path:
    fm = parse_frontmatter(md_path)
    doc = (fm.get("doc") or "").strip()
    tmpl = TEMPLATE_BY_DOC.get(doc)
    if tmpl is None:
        raise ValueError(
            f"doc: '{doc}' が不正です。{list(TEMPLATE_BY_DOC)} のいずれかにしてください"
        )
    if not fm.get("client"):
        raise ValueError("client(宛先)が空です")

    wb = openpyxl.load_workbook(tmpl)
    ws = wb.active
    if doc == "御請求書":
        fill_invoice(ws, fm)
    else:
        fill_estimate(ws, fm)

    if out_path is None:
        out_path = md_path.with_suffix(".xlsx")

    if dry_run:
        # 検算: 明細の税抜合計だけ Python 側でも出す(雛形数式の答え合わせ用)
        items = fm.get("line_items", [])
        subtotal = sum(it["unit_price"] * it.get("qty", 1) for it in items)
        tax = int(subtotal * 0.1)
        print(f"[dry-run] doc={doc} client={fm['client']} 明細={len(items)}行")
        print(f"[dry-run] 税抜小計={subtotal:,} 消費税(目安)={tax:,} 税込={subtotal+tax:,}")
        print(f"[dry-run] 出力先(未保存)= {out_path}")
        return out_path

    wb.save(out_path)
    print(f"発行しました: {out_path}")
    return out_path


def main():
    ap = argparse.ArgumentParser(description="構造正本 md → 雛形 .xlsx 発行")
    ap.add_argument("md", help="入力の構造正本 md(frontmatter 必須)")
    ap.add_argument("--out", help="出力 .xlsx パス(既定: md と同じ場所)")
    ap.add_argument("--dry-run", action="store_true", help="保存せず検算のみ")
    args = ap.parse_args()

    md_path = Path(args.md).resolve()
    if not md_path.exists():
        print(f"入力が見つかりません: {md_path}", file=sys.stderr)
        sys.exit(1)
    out = Path(args.out).resolve() if args.out else None
    issue(md_path, out, args.dry_run)


if __name__ == "__main__":
    main()
